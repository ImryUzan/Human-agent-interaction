from __future__ import print_function
from gevent import os
from ortools.sat.python import cp_model
import openpyxl

def main():
    num_employee = int(input("Please enter the number of employees:\n"))
    num_shifts = int(input("Please enter the number of shifts:\n"))
    num_employees_per_shift = []
    print("Please enter the number of employees of shifts:")
    for i in range(num_shifts):
        num_employees_per_shift.append(int(input()))
    num_days = int(input("Please enter the number of days:\n"))
    allocation_method = int(input("if you want to consider fairness press 2, otherwise press 1:\n"))
    all_employees = range(num_employee)
    all_shifts = range(num_shifts)
    all_days = range(num_days)
    # Give the location of the file (assumes the xlsx file is in project dir and called Requests.xlsx)
    root = os.path.dirname(os.path.abspath(__file__))
    path = root + "\\Requests.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    # all the days employee wants to work
    listOfPossibleDays = []
    listOfDicts = []
    for i in all_employees:
        col_num = 2
        listOfDicts.append({})
        cell_days = sheet_obj.cell(row=i + 2, column=num_shifts * num_days + col_num)
        listOfPossibleDays.append(cell_days.value)
        for j in range(num_days):
            for g in range(num_shifts):
                # list of shifts of 1 employee per 1 day
                cell_obj = sheet_obj.cell(row=i+2, column=col_num)
                val = 0
                if cell_obj.value == 1:
                    val = -5
                elif cell_obj.value == 2:
                    val = -2
                elif cell_obj.value == 3:
                    val = 0
                elif cell_obj.value == 4:
                    val = 4
                elif cell_obj.value == 5:
                    val = 5
                listOfDicts[i][chr(ord('A') + g) + str(j + 1)] = val
                col_num = col_num+1
    max_value = 5
    shift_req_day_t = []
    shift_req_week_t = []
    shift_requests = []
    for row in listOfDicts:
        for j in range(num_days):
            for g in range(num_shifts):
                # list of shifts of 1 employee per 1 day
                shift_req_day_t.append(row[chr(ord('A') + g) + str(j + 1)])
            shift_req_day = shift_req_day_t[:]
            # saves employee's shifts as list of lists ,each inner list is shifts per day ,outter list presents all week
            shift_req_week_t.append(shift_req_day)
            shift_req_day_t.clear()
        shift_req_week = shift_req_week_t[:]
        # list of all weekly lists of all the employee's (list of lists of lists)
        shift_requests.append(shift_req_week)
        shift_req_week_t.clear()

    model = cp_model.CpModel()
    shifts = {}
    for n in all_employees:
        for d in all_days:
            for s in all_shifts:
                shifts[(n, d, s)] = model.NewBoolVar('shift_n%id%is%i' % (n, d, s))

    for d in all_days:
        for e in range(len(num_employees_per_shift)):
            # number of employee's per shift equals to num_employees_per_shift[e]
            model.Add(sum(shifts[(n, d, e)] for n in all_employees) == num_employees_per_shift[e])

    for n in all_employees:
        model.Add(sum(shifts[(n, d, s)] for d in all_days for s in all_shifts) <= listOfPossibleDays[n])
        for d in all_days:
            # 1 shift per day for employee
            model.Add(sum(shifts[(n, d, s)] for s in all_shifts) <= 1)

    if allocation_method == 1:
        model.Maximize(sum(
            shift_requests[n][d][s] * shifts[(n, d, s)] for n in all_employees for d in all_days for s in all_shifts))
        solver = cp_model.CpSolver()
        solver.Solve(model)
        best_solver = solver

    if allocation_method == 2:
        model.Maximize(sum(
            shift_requests[n][d][s] * shifts[(n, d, s)] for n in all_employees for d in all_days for s in all_shifts))
        solver = cp_model.CpSolver()
        solver.Solve(model)
        best_solver = solver
        for i in range(10):
            for n in all_employees:
                model.Add((sum(shift_requests[n][d][s] * shifts[(n, d, s)]
                               for d in all_days for s in all_shifts))*10 >= i*listOfPossibleDays[n]*max_value)
            model.Maximize(sum(
                shift_requests[n][d][s] * shifts[(n, d, s)] for n in all_employees for d in all_days for s in
                all_shifts))
            solver = cp_model.CpSolver()
            solver.Solve(model)
            solver_status = str(solver.ResponseStats())
            if "INFEASIBLE" in solver_status:
                break
            else:
                best_solver = solver

    assign_employees = ""
    counters = [0] * num_employee
    for d in all_days:
        for s in all_shifts:
            for n in all_employees:
                if best_solver.Value(shifts[(n, d, s)]) == 1:
                    assign_employees += str(n + 1) + ", "
                    counters[n] += shift_requests[n][d][s] /(listOfPossibleDays[n] * max_value)
            print('Shift ' + str(s + 1) + ", " + 'Day ' + str(d + 1)+ ": " + assign_employees[:-2])

            assign_employees = ""

    # Statistics.
    print()
    print('Statistics')
    print('  - Employees satisfaction status :')
    Total_percentage = []
    for i in range(len(counters)):
        percentage = counters[i] * 100
        print('     Employee ' + str(i + 1) + ' is ' + str(percentage) + '% Statisfied')
        Total_percentage.append(percentage)


if __name__ == '__main__':
    main()